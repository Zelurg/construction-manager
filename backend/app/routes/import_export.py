from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import date, datetime
from .. import models, schemas
from ..database import get_db
from ..dependencies import get_current_user
from .projects import touch_project

router = APIRouter()

COLUMN_MAP = {
    'A': 'code', 'B': 'name', 'C': 'unit',
    'D': 'volume_plan', 'E': 'start_date_contract', 'F': 'end_date_contract',
    'G': 'start_date_plan', 'H': 'end_date_plan',
    'I': 'unit_price', 'J': 'labor_per_unit', 'K': 'machine_hours_per_unit',
    'L': 'executor', 'M': 'status_people', 'N': 'status_equipment',
    'O': 'status_mtr', 'P': 'status_access',
}


def parse_float(v):
    if v is None or str(v).strip() in ('', '-', 'None'):
        return 0.0
    try:
        return float(str(v).replace(',', '.'))
    except:
        return 0.0


def parse_date(v):
    if v is None:
        return None
    if hasattr(v, 'date'):
        return v.date()
    from datetime import datetime as dt
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return dt.strptime(str(v).strip(), fmt).date()
        except:
            pass
    return None


def parse_status(v):
    if v is None or str(v).strip() == '':
        return 'gray'
    val = str(v).strip().lower()
    if val in ('green', 'зелёный', 'зеленый'):
        return 'green'
    elif val in ('yellow', 'жёлтый', 'желтый'):
        return 'yellow'
    elif val in ('red', 'красный'):
        return 'red'
    return 'gray'


def status_to_text(status):
    mapping = {'green': 'Зелёный', 'yellow': 'Жёлтый', 'red': 'Красный', 'gray': ''}
    return mapping.get(status, '')


@router.post("/import")
async def import_tasks(
    file: UploadFile = File(...),
    project_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Импорт графика с обновлением существующих задач по коду"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Поддерживаются только файлы .xlsx и .xls")

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    tasks_created = 0
    tasks_updated = 0
    errors = []

    existing_tasks = {}
    if project_id:
        for task in db.query(models.Task).filter(models.Task.project_id == project_id).all():
            existing_tasks[task.code] = task

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            task_data = {}
            cols = list('ABCDEFGHIJKLMNOP')
            for i, col in enumerate(cols):
                if i < len(row):
                    task_data[COLUMN_MAP[col]] = row[i]

            code = str(task_data.get('code', '')).strip() if task_data.get('code') else None
            name = str(task_data.get('name', '')).strip() if task_data.get('name') else None
            if not code or not name:
                errors.append(f"Строка {row_num}: пропущен код или название")
                continue

            raw_name = task_data.get('name', '')
            if isinstance(raw_name, str):
                spaces = len(raw_name) - len(raw_name.lstrip())
                level = spaces // 2
            else:
                level = 0

            unit = task_data.get('unit')
            is_section = not unit or str(unit).strip() == ''

            start_contract = parse_date(task_data.get('start_date_contract'))
            end_contract   = parse_date(task_data.get('end_date_contract'))
            start_plan     = parse_date(task_data.get('start_date_plan'))
            end_plan       = parse_date(task_data.get('end_date_plan'))

            if start_plan is None:
                start_plan = start_contract
            if end_plan is None:
                end_plan = end_contract

            update_data = {
                "name": name.strip(),
                "unit": str(unit).strip() if unit and str(unit).strip() else None,
                "volume_plan": parse_float(task_data.get('volume_plan')),
                "start_date_contract": start_contract,
                "end_date_contract": end_contract,
                "start_date_plan": start_plan,
                "end_date_plan": end_plan,
                "unit_price": parse_float(task_data.get('unit_price')),
                "labor_per_unit": parse_float(task_data.get('labor_per_unit')),
                "machine_hours_per_unit": parse_float(task_data.get('machine_hours_per_unit')),
                "executor": str(task_data.get('executor', '')).strip() or None,
                "is_section": is_section,
                "level": level,
                "status_people": parse_status(task_data.get('status_people')),
                "status_equipment": parse_status(task_data.get('status_equipment')),
                "status_mtr": parse_status(task_data.get('status_mtr')),
                "status_access": parse_status(task_data.get('status_access')),
            }

            if code in existing_tasks:
                task = existing_tasks[code]
                for key, value in update_data.items():
                    setattr(task, key, value)
                tasks_updated += 1
            else:
                new_task = models.Task(
                    project_id=project_id,
                    code=code,
                    volume_fact=0.0,
                    sort_order=(row_num - 1) * 10,
                    **update_data
                )
                db.add(new_task)
                existing_tasks[code] = new_task
                tasks_created += 1

        except Exception as e:
            errors.append(f"Строка {row_num}: {str(e)}")

    stack = []
    for code in existing_tasks.keys():
        task = existing_tasks[code]
        while stack and stack[-1]['level'] >= task.level:
            stack.pop()
        task.parent_code = stack[-1]['code'] if stack else None
        if task.is_section:
            stack.append({'code': task.code, 'level': task.level})

    db.commit()
    touch_project(project_id, db)

    return {
        "tasks_created": tasks_created,
        "tasks_updated": tasks_updated,
        "errors": errors
    }


@router.get("/export")
def export_tasks(
    project_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """Экспорт графика со всеми атрибутами"""
    query = db.query(models.Task)
    if project_id is not None:
        query = query.filter(models.Task.project_id == project_id)
    tasks = query.order_by(models.Task.sort_order, models.Task.code).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График"

    headers = [
        'Код', 'Наименование', 'Ед.изм.', 'Объём план',
        'Нач.контракт', 'Оконч.контракт', 'Нач.план', 'Оконч.план',
        'Цена за ед.', 'Трудозатраты/ед.', 'Машиночасы/ед.', 'Исполнитель',
        'Люди', 'Техника', 'МТР', 'Допуск'
    ]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_num, task in enumerate(tasks, 2):
        indent = '  ' * task.level if task.level else ''
        ws.cell(row=row_num, column=1, value=task.code)
        ws.cell(row=row_num, column=2, value=f"{indent}{task.name}")
        ws.cell(row=row_num, column=3, value=task.unit)
        ws.cell(row=row_num, column=4, value=task.volume_plan)
        ws.cell(row=row_num, column=5, value=str(task.start_date_contract) if task.start_date_contract else None)
        ws.cell(row=row_num, column=6, value=str(task.end_date_contract) if task.end_date_contract else None)
        ws.cell(row=row_num, column=7, value=str(task.start_date_plan) if task.start_date_plan else None)
        ws.cell(row=row_num, column=8, value=str(task.end_date_plan) if task.end_date_plan else None)
        ws.cell(row=row_num, column=9, value=task.unit_price)
        ws.cell(row=row_num, column=10, value=task.labor_per_unit)
        ws.cell(row=row_num, column=11, value=task.machine_hours_per_unit)
        ws.cell(row=row_num, column=12, value=task.executor)
        ws.cell(row=row_num, column=13, value=status_to_text(task.status_people))
        ws.cell(row=row_num, column=14, value=status_to_text(task.status_equipment))
        ws.cell(row=row_num, column=15, value=status_to_text(task.status_mtr))
        ws.cell(row=row_num, column=16, value=status_to_text(task.status_access))

        if task.is_section:
            for col in range(1, 17):
                ws.cell(row=row_num, column=col).font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=schedule_export.xlsx"}
    )


@router.get("/export-msg")
def export_msg(
    project_id: int = Query(...),
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db)
):
    """Экспорт МСГ: только работы выбранного месяца + их родительские секции, без колонок дней"""
    from calendar import monthrange

    month_start = date(year, month, 1)
    _, days_in_month = monthrange(year, month)
    month_end = date(year, month, days_in_month)

    all_tasks = db.query(models.Task).filter(
        models.Task.project_id == project_id
    ).order_by(models.Task.sort_order, models.Task.code).all()

    if not all_tasks:
        raise HTTPException(status_code=404, detail="Задачи не найдены")

    work_ids_in_month = set()
    for task in all_tasks:
        if task.is_section:
            continue
        s = task.start_date_plan
        e = task.end_date_plan
        if task.is_custom and not s and not e:
            work_ids_in_month.add(task.id)
            continue
        if s and e and s <= month_end and e >= month_start:
            work_ids_in_month.add(task.id)

    section_codes_needed = set()
    for task in all_tasks:
        if task.id not in work_ids_in_month:
            continue
        parts = str(task.code).split('.')
        for length in range(1, len(parts)):
            section_codes_needed.add('.'.join(parts[:length]))

    visible_tasks = [
        t for t in all_tasks
        if t.id in work_ids_in_month or (t.is_section and t.code in section_codes_needed)
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"МСГ {year}-{month:02d}"

    headers = [
        'Код', 'Наименование', 'Ед.изм.', 'Объём план',
        'Нач.контракт', 'Оконч.контракт', 'Нач.план', 'Оконч.план',
        'Цена за ед.', 'Трудозатраты/ед.', 'Машиночасы/ед.', 'Исполнитель',
        'Люди', 'Техника', 'МТР', 'Допуск'
    ]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    SECTION_COLORS = ['B8D4E8', 'C8DFF0', 'D8EAF5', 'E4F1F8', 'EFF6FB']

    for row_num, task in enumerate(visible_tasks, 2):
        indent = '  ' * task.level if task.level else ''
        ws.cell(row=row_num, column=1, value=task.code)
        ws.cell(row=row_num, column=2, value=f"{indent}{task.name}")
        ws.cell(row=row_num, column=3, value=task.unit)
        ws.cell(row=row_num, column=4, value=task.volume_plan)
        ws.cell(row=row_num, column=5, value=str(task.start_date_contract) if task.start_date_contract else None)
        ws.cell(row=row_num, column=6, value=str(task.end_date_contract) if task.end_date_contract else None)
        ws.cell(row=row_num, column=7, value=str(task.start_date_plan) if task.start_date_plan else None)
        ws.cell(row=row_num, column=8, value=str(task.end_date_plan) if task.end_date_plan else None)
        ws.cell(row=row_num, column=9, value=task.unit_price)
        ws.cell(row=row_num, column=10, value=task.labor_per_unit)
        ws.cell(row=row_num, column=11, value=task.machine_hours_per_unit)
        ws.cell(row=row_num, column=12, value=task.executor)
        ws.cell(row=row_num, column=13, value=status_to_text(task.status_people))
        ws.cell(row=row_num, column=14, value=status_to_text(task.status_equipment))
        ws.cell(row=row_num, column=15, value=status_to_text(task.status_mtr))
        ws.cell(row=row_num, column=16, value=status_to_text(task.status_access))

        if task.is_section:
            level_idx = min(task.level or 0, len(SECTION_COLORS) - 1)
            fill_color = SECTION_COLORS[level_idx]
            section_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col in range(1, 17):
                cell = ws.cell(row=row_num, column=col)
                cell.font = Font(bold=True)
                cell.fill = section_fill

    col_widths = [12, 50, 8, 12, 14, 14, 12, 12, 12, 16, 16, 20, 10, 10, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=msg_{year}_{month:02d}.xlsx"}
    )


@router.post("/import-msg")
async def import_msg(
    file: UploadFile = File(...),
    project_id: int = Query(...),
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Импорт МСГ: читаем те же 16 колонок что экспорт МСГ и обновляем поля задач.
    Новые задачи НЕ создаются — только обновление существующих.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Поддерживаются только файлы .xlsx и .xls")

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    # Индекс всех задач проекта по коду
    tasks = db.query(models.Task).filter(
        models.Task.project_id == project_id
    ).all()
    task_map = {task.code: task for task in tasks}

    tasks_updated = 0
    tasks_skipped = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        try:
            # Колонка A — код задачи
            code = str(row[0]).strip() if row[0] is not None else None
            if not code:
                continue

            if code not in task_map:
                tasks_skipped += 1
                errors.append(f"Строка {row_num}: код '{code}' не найден в проекте")
                continue

            task = task_map[code]

            # Читаем колонки (0-based индекс в row):
            # 0=код, 1=название, 2=ед, 3=объём,
            # 4=нач_контр, 5=оконч_контр, 6=нач_план, 7=оконч_план,
            # 8=цена, 9=трудоз, 10=машчас, 11=исполнитель,
            # 12=люди, 13=техника, 14=мтр, 15=допуск

            def get_col(idx):
                return row[idx] if idx < len(row) else None

            # Обновляем только непустые поля
            unit = get_col(2)
            if unit is not None:
                task.unit = str(unit).strip() or None

            vol = get_col(3)
            if vol is not None:
                task.volume_plan = parse_float(vol)

            sc = parse_date(get_col(4))
            if sc is not None:
                task.start_date_contract = sc

            ec = parse_date(get_col(5))
            if ec is not None:
                task.end_date_contract = ec

            sp = parse_date(get_col(6))
            if sp is not None:
                task.start_date_plan = sp

            ep = parse_date(get_col(7))
            if ep is not None:
                task.end_date_plan = ep

            price = get_col(8)
            if price is not None:
                task.unit_price = parse_float(price)

            labor = get_col(9)
            if labor is not None:
                task.labor_per_unit = parse_float(labor)

            mach = get_col(10)
            if mach is not None:
                task.machine_hours_per_unit = parse_float(mach)

            executor = get_col(11)
            if executor is not None:
                task.executor = str(executor).strip() or None

            sp_people = get_col(12)
            if sp_people is not None:
                task.status_people = parse_status(sp_people)

            sp_equip = get_col(13)
            if sp_equip is not None:
                task.status_equipment = parse_status(sp_equip)

            sp_mtr = get_col(14)
            if sp_mtr is not None:
                task.status_mtr = parse_status(sp_mtr)

            sp_access = get_col(15)
            if sp_access is not None:
                task.status_access = parse_status(sp_access)

            tasks_updated += 1

        except Exception as e:
            errors.append(f"Строка {row_num}: {str(e)}")

    db.commit()
    touch_project(project_id, db)

    return {
        "tasks_updated": tasks_updated,
        "tasks_skipped": tasks_skipped,
        "errors": errors
    }
